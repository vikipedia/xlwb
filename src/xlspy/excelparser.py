import sys, time, argparse, os
import re
import collections
import functools
import operator
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from excelfunctions import functionsmap, OFFSET
from debug import debugmethods, trace
from memoize import memoize
import tree_evaluator
import pickle

Token = collections.namedtuple("Token",['type','subtype','value'])

def generate_tokens(text):
    """
    token generator. Uses openpyxl.formula.Tokenizer to get tokens
    """
    tok = Tokenizer(text)
    for item in tok.items:
        token = Token(item.type, item.subtype, item.value)
        if token.type != "WHITE-SPACE":
            yield token

#Utility functions
def get_cell(w, sheetname,column,row):
    s = w[sheetname]
    c = s["".join([column,row])]
    return c


def pretty_print(tree):
    def print_(tree, level):
        op, args = tree[0], tree[1:]
        print(" "*level + op)
        for a in args:
            if isinstance(a, tuple):
                print_(a, level+1)
            else:
                print(" "*(level+1) + str(a))
                      
    print_(tree, 0)



class ExpressionEvaluator:

    def __init__(self, workbook=None):
        self.workbook = workbook

    @classmethod
    def instance(cls, *args):
        return cls(*args)
        
    
    def parse(self, text):
        self.tokens = generate_tokens(text)
        self.tok = None #last token consumed
        self.nexttok = None #Next symbol to be tokenized
        self._advance()
        return self.expr()

    def _advance(self):
        "Advance one token ahead"
        self.tok, self.nexttok = self.nexttok, next(self.tokens, None)

    def _accept(self, toktype, assert_=lambda :True):
        "Test and consume the next token of it matches toktype"
        if self.nexttok and self.nexttok.type == toktype and assert_():
            self._advance()
            return True
        else:
            return False
        

    def _expect(self,toktype, check=lambda:True):
        'Consume next token if it matches toktype or raise SyntaxError'
        
        if not self._accept(toktype) and check():
            raise SyntaxError('Expected ' + toktype)


    # Grammar rules follow
    def expr(self):
        "expression ::= expr { ('='|'<'|'>'|'<='|'>=' | '<>' ) expr }*"
        exprval = self.strexpr() 
        assert_ = lambda :self.nexttok.value in ["=",">","<","<=",">=", "<>"]

        while self._accept('OPERATOR-INFIX',assert_):
            op = self.tok.value
            left = exprval or 0
            right = self.strexpr() or 0 #FIXME what is it returns string?
            exprval = self.create_node(op, left, right)
        return exprval

    def strexpr(self):
        "string expression ::= expr_ { & } expr_"
        exprval = self.expr_() 
        assert_ = lambda :self.nexttok.value in ["&"]
        
        while self._accept('OPERATOR-INFIX',assert_):
            op = self.tok.value
            left = exprval or ""
            right = self.expr_() or ""
            exprval = self.create_node(op, left, right)
        return exprval
    
    def expr_(self):
        "expression ::= term { ('+'|'-') term }*"
        exprval = self.term()
        assert_ = lambda :self.nexttok.value in ["+","-"]

        while  self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            left = exprval or 0
            right = self.term() or 0
            exprval = self.create_node(op, left, right)
        return exprval

    
    def term(self):
        "term ::= factor { ('*'|'/') factor }*"
        termval = self.power()
        assert_ = lambda :self.nexttok.value in ["*","/"]

        while self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            left = termval
            right = self.power()
            termval = self.create_node(op, left, right)
        return termval


    def power(self):
        "power ::= percent ^ n "
        termval = self.percent()
        assert_ = lambda :self.nexttok.value == "^"

        while self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            left = termval
            right = self.percent()
            termval = self.create_node(op, left, right)
        return termval


    def percent(self):
        "modulus = factor {%}"
        termval = self.factor()
        assert_ = lambda : self.nexttok.value == "%"

        while self._accept("OPERATOR-POSTFIX", assert_):
            op = self.tok.value
            left = termval
            termval = self.create_node("/", left, 100)
        return termval

    def negation(self, pre):
        v = self.factor()
        return self.create_node("*", pre, v)
    
    def factor(self):
        "factor ::= FUNC | RANGE| NUM | ( expr )"
        assert_L = lambda : self.nexttok.subtype == "LOGICAL"
        assert_T = lambda : self.nexttok.subtype == "TEXT"
        assert_N = lambda : self.nexttok.subtype == "NUMBER"
        assert_R = lambda : self.nexttok.subtype == "RANGE"
        assert_O = lambda : self.nexttok.subtype == "OPEN"
        assert_E = lambda : self.nexttok.subtype == "ERROR"
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"

        assert_ = lambda : self.nexttok.value in ["-","+"]
        if self._accept("OPERATOR-PREFIX", assert_):
            return self.negation({"-":-1,"+":1}[self.tok.value])
        elif self._accept("FUNC", assert_O):
            return self.function()
        elif self._accept('OPERAND', assert_R):
            return self.range(self.tok.value)
        elif self._accept('OPERAND', assert_N):
            return float(self.tok.value)
        elif self._accept('OPERAND', assert_L):
            return False if self.tok.value=="FALSE" else True
        elif self._accept('OPERAND', assert_T):
            return self.text_(self.tok.value)
        elif self._accept("LITERAL"):
            return self.tok.value
        elif self._accept("OPERAND",assert_E):
            return None
        elif self._accept('PAREN',assert_O):
            exprval = self.expr()
            self._expect('PAREN', assert_C)
            return exprval
        else:
            raise SyntaxError('Expected NUMBER or LPAREN')


    def text_(self, text):
        if text.startswith("'") and text.endswith("'"):
            return text.strip().replace("'","").strip()
        elif text.startswith('"') and text.endswith('"'):
            return text.strip().replace('"',"").strip()
        else:
            return text.strip()
        
    def range(self, text):
        """
        evaluate excel cell ranges
        """
        namedranges = [r.name for r in self.workbook.get_named_ranges()]
        if text in namedranges:
            namedrange = self.workbook.get_named_range(text)
            text = namedrange.attr_text
        
        if ":" in text:
            return self.range_(text)
        else:
            return self.cell(text)


        
    def cell(self, text):
        """
        evaluate indivudual cell of excel
        """
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<COL>[A-Z]+)\$?(?P<ROW>\d+)")
        m = pattern.match(text)
        if pattern.match(text):
            return self.cell1(text, pattern)
        else:
            return self.cell1_(text)


    def cell1_(self, text):
        """
        evaluate cells on active sheet
        """

        sheet = self.workbook.active
        c = sheet[self.tok.value]
        return self.parsecell(c)

    
    def cell1(self, text, pattern):
        """
        evaluate cells from non active sheet
        """

        m = pattern.match(text)
        sheet, col, row = m.groups()
        c = get_cell(self.workbook, sheet, col, row)
        active_ = self.workbook.active.title
        self.workbook.active = self.workbook.get_sheet_names().index(sheet)
        v = self.parsecell(c)
        self.workbook.active = self.workbook.get_sheet_names().index(active_)
        return v
    
    def parsecell(self, c):
        address =  "!".join([self.workbook.get_active_sheet().title,c.coordinate])
        if c.data_type == c.TYPE_FORMULA:
                return ("CELL", address)
        else: 
            return c.value

    def range_value(self, textattr, pattern):
        m = pattern.match(textattr)
    
        sheet, ranges = m.groups()
        active_ = self.workbook.active.title
        self.workbook.active = self.workbook.get_sheet_names().index(sheet)
        s = self.workbook.active
        v = [[self.parsecell(col) for col in row] for row in s[ranges]]
        self.workbook.active = self.workbook.get_sheet_names().index(active_)
        return v
        
    def range_(self, text):
        
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<RANGE>[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)")
        m = pattern.match(text)

        if m:
            return self.range_value(text, pattern)
        else:
            s = self.workbook.active
            v= [[self.parsecell(col) for col in row] for row in s[text]]

            return v
            #return [self.parsecell(c[0]) for c in s[self.tok.value]]

    def create_node(self, op,  *args):
        if not args:
            return op
        else:
            return functionsmap[op](*args)

            
            
    def OFFSET(self):
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"
    
        self._advance()
        ref = self.tok.value # take this as literal string
        
        args = []
        while self._accept("SEP", assert_A):
            args.append(self.expr())
        self._expect("FUNC", assert_C)

        try:
            args = [tree_evaluator.evaluate(a, {}) for a in args]
            r = OFFSET(ref, *args)
        except Exception as e:
            print("OFFSET Failure!", ref, args)
            return None
        return self.range(r)
        
    def function(self):
        """
        function ::=  FUNC (EXPR,EXPR..)
        """
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"
    
        funcname,_ = self.tok.value.split("(")
        if funcname == "OFFSET":
            return self.OFFSET()
        
        args = [self.expr()]
        while self._accept("SEP", assert_A):
            args.append(self.expr())
        self._expect("FUNC", assert_C)
        return self.create_node(funcname, *args)

   
class ExpressionTreeBuilder(ExpressionEvaluator):


    def create_node(self, op, *args):
        if not args:
            return op
        else:
            return (op, *args)

    def parsecell(self, c):
        address =  "!".join([self.workbook.get_active_sheet().title,c.coordinate])
        return ("CELL", address)
        

def create_cellmap(filename, inputs):
    w = load_workbook(filename=filename)
    cellmap = {}
    
    et = ExpressionTreeBuilder(w)
    
    
    #put named ranges in cellmap
    for r in w.get_named_ranges():
        try:
            e = et.parse("="+r.attr_text)
            cellmap[r.name] = e
        except Exception as e:
            print("Skipping ", r.name , e)

    #put every cell in cellmap
    for i,name in enumerate(w.get_sheet_names()):
        w.active = i
        sheet = w[name]
        for col in sheet:
            for c in col:
                cellid = "!".join([name, c.coordinate])
                if c.value==None:
                    pass
                else:
                    if c.data_type == c.TYPE_FORMULA:
                        e = et.parse(c.value)
                        cellmap[cellid] = e
                    else:
                        cellmap[cellid] = c.value
                        
    cellmap.update(inputs)
    return cellmap
    
def output_extn(filename):
    tokens = filename.split(".")
    return ".".join(tokens[:-1] + ["bin"])


def parse_args():
    parser = argparse.ArgumentParser("Excel data generator")
    parser.add_argument("filename",
                        type=str,
                        help="input excel filename")

    parser.add_argument("-o","--output",
                        type=str,
                        help="output filename, if not given input file's name will be used with bin extension.")
    return parser.parse_args()


def main(filename, output=None):
    if not output:
        output = output_extn(os.path.basename(filename))
    cm = create_cellmap(filename, {})
    
    with open(output, "wb") as o:
        pickle.dump(cm, o)
        

if __name__ == "__main__":
    args = parse_args()
    main(args.filename, args.output)
    
