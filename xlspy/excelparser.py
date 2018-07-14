import sys
import re
import collections
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from excelfunctions import functionsmap
from debug import debugmethods

Token = collections.namedtuple("Token",['type','subtype','value'])

def generate_tokens(text):
    """
    token generator. Uses openpyxl.formula.Tokenizer to get tokens
    """
    tok = Tokenizer(text)
    for item in tok.items:
        token = Token(item.type, item.subtype, item.value)
        if token.type != "WHITE-SPACE":
            yield token

#Utility functions
def get_cell(w, sheetname,column,row):
    s = w.get_sheet_by_name(sheetname)
    c = s.cell(column = columnind(column), row=int(row))
    return c

def columnind(ch):
    """
    accepts column numbers in excels' columnname format
    and returns actual column number.
    >>> column("A")
    1
    >>> column("AA")
    27
    """
    n = len(ch)
    return sum((ord(c)-64)*26**(n-i-1) for i,c in enumerate(ch))

def extract_column_row(cell):
    pattern = re.compile(r'\$?(?P<COL>[A-Z]+)\$?(?P<ROW>\d+)')
    m = pattern.match(cell)
    column, row = m.groups()
    return column, row


def pretty_print(tree):
    def print_(tree, level):
        op, args = tree[0], tree[1:]
        print(" "*level + op)
        for a in args:
            if isinstance(a, tuple):
                print_(a, level+1)
            else:
                print(" "*(level+1) + str(a))
                      
    print_(tree, 0)

def flatten(items):
    fl = []
    for item in items:
        if isinstance(item, list):
            fl.extend(item)
        else:
            fl.append(item)
    return fl

#@debugmethods
class ExpressionEvaluator:

    def __init__(self, workbook=None):
        self.workbook = workbook
    
    def parse(self, text):
        self.tokens = generate_tokens(text)
        self.tok = None #last token consumed
        self.nexttok = None #Next symbol to be tokenized
        self._advance()
        return self.expr()

    def _advance(self):
        "Advance one token ahead"
        self.tok, self.nexttok = self.nexttok, next(self.tokens, None)

    def _accept(self, toktype, assert_=lambda :True):
        "Test and consume the next token of it matches toktype"
        if self.nexttok and self.nexttok.type == toktype and assert_():
            self._advance()
            return True
        else:
            return False
        

    def _expect(self,toktype, check=lambda:True):
        'Consume next token if it matches toktype or raise SyntaxError'
        
        if not self._accept(toktype) and check():
            raise SyntaxError('Expected ' + toktype)


    # Grammar rules follow
    def expr(self):
        "expression ::= expr { ('='|'<'|'>'|'<='|'>=') expr }*"
        exprval = self.expr_()
        assert_ = lambda :self.nexttok.value in ["=",">","<","<=",">="]
        while self._accept('OPERATOR-INFIX',assert_):
            op = self.tok.value
            right = self.expr_()
            if op == "=":
                exprval = exprval == right
            elif op == '<':
                exprval = exprval < right
            elif op == '>':
                exprval = exprval > right
            elif op == ">=":
                exprval = exprval >= right
            elif op == "<=":
                exprval = exprval <= right
        return exprval        
    
    
    def expr_(self):
        "expression ::= term { ('+'|'-') term }*"
        exprval = self.term()
        assert_ = lambda :self.nexttok.value in ["+","-"]
        while  self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            left = exprval or 0
            right = self.term() or 0
            if op == '+':
                exprval = left + right
            elif op == '-':
                exprval = left - right
        return exprval

    
    def term(self):
        "term ::= factor { ('*'|'/') factor }*"
        termval = self.power()
        assert_ = lambda :self.nexttok.value in ["*","/"]
        while self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            right = self.power()
            if op == '*':
                termval *= right
            elif op == '/':
                termval /= right
        return termval


    def power(self):
        "power ::= percent ^ n "
        termval = self.percent()
        assert_ = lambda :self.nexttok.value == "^"
        while self._accept('OPERATOR-INFIX', assert_):
            op = self.tok.value
            left = termval
            right = self.percent()
            termval = left ** right
        return termval


    def percent(self):
        "modulus = factor {%}"
        termval = self.factor()
        assert_ = lambda : self.nexttok.value == "%"
        while self._accept("OPERATOR-POSTFIX", assert_):
            op = self.tok.value
            left = termval
            termval = left/100
        return termval

    def negation(self, pre):
        v = self.factor()
        return pre*v
    
    def factor(self):
        "factor ::= FUNC | RANGE| NUM | ( expr )"
        assert_T = lambda : self.nexttok.subtype == "TEXT"
        assert_N = lambda : self.nexttok.subtype == "NUMBER"
        assert_R = lambda : self.nexttok.subtype == "RANGE"
        assert_O = lambda : self.nexttok.subtype == "OPEN"
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"


        assert_ = lambda : self.nexttok.value in ["-","+"]
        if self._accept("OPERATOR-PREFIX", assert_):
            return self.negation({"-":-1,"+":1}[self.tok.value])
        elif self._accept("FUNC", assert_O):
            return self.function()
        elif self._accept('OPERAND', assert_R):
            return self.range()
        elif self._accept('OPERAND', assert_N):
            return float(self.tok.value)
        elif self._accept('OPERAND', assert_T):
            return self.tok.value
        elif self._accept("LITERAL"):
            return self.tok.value
        elif self._accept('PAREN',assert_O):
            exprval = self.expr()
            self._expect('PAREN', assert_C)
            return exprval
        else:
            raise SyntaxError('Expected NUMBER or LPAREN')


        
    def range(self):
        """
        evaluate excel cell ranges
        """
        namedranges = [r.name for r in self.workbook.get_named_ranges()]
        if self.tok.value in namedranges:
            return self.namedrange_(self.tok.value)
        elif ":" in self.tok.value:
            return self.range_()
        else:
            return self.cell()

    def namedrange_(self, rangename):
        r = self.workbook.get_named_range(rangename)
        

    def cell(self):
        """
        evaluate indivudual cell of excel
        """
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<COL>[A-Z]+)\$?(?P<ROW>\d+)")
        m = pattern.match(self.tok.value)
        if pattern.match(self.tok.value):
            return self.cell1()
        else:
            return self.cell1_()

    def cell1_(self):
        """
        evaluate cells on active sheet
        """

        col, row = extract_column_row(self.tok.value)
        sheet = self.workbook.active
        c = sheet[self.tok.value]
        return self.parsecell(c)

    def cell1(self):
        """
        evaluate cells from non active sheet
        """
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<COL>[A-Z]+)\$?(?P<ROW>\d+)")
        m = pattern.match(self.tok.value)
        sheet, col, row = m.groups()
        c = get_cell(self.workbook, sheet, col, row)
        active_ = self.workbook.active.title
        self.workbook.active = self.workbook.get_sheet_names().index(sheet)
        v = self.parsecell(c)
        self.workbook.active = self.workbook.get_sheet_names().index(active_)
        return v
    
    def parsecell(self, c):
        if c.data_type == c.TYPE_FORMULA:
            e = ExpressionEvaluator(self.workbook)
            return e.parse(c.value)
        else: 
            return c.value


    def range_value(self, textattr):
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<RANGE>[A-Z]+\d+:[A-Z]+\d+)")
        m = pattern.match(textattr)
    
        sheet, ranges = m.groups()
        active_ = self.workbook.active.title
        self.workbook.active = self.workbook.get_sheet_names().index(sheet)
        s = self.workbook.active
        v = [self.parsecell(c[0]) for c in s[ranges]]
        self.workbook.active = self.workbook.get_sheet_names().index(active_)
        return v            
        
        
    def range_(self):
        pattern = re.compile(r"\$?'?(?P<SHEET>[\w &-]+)'?[\!\.]\$?(?P<RANGE>[A-Z]+\d+:[A-Z]+\d+)")
        m = pattern.match(self.tok.value)
        if m:
            return self.range_value(self.tok.value)
        else:
            s = self.workbook.active
            v = []
            v= [[self.parsecell(c) for c in items]for items in s[self.tok.value]]

            return flatten(v)          
            #return [self.parsecell(c[0]) for c in s[self.tok.value]]


    def function(self):
        """
        function ::=  FUNC (EXPR,EXPR..)
        """
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"
    
        funcname,_ = self.tok.value.split("(")
        args = [self.expr()]
        while self._accept("SEP", assert_A):
            args.append(self.expr())
        self._expect("FUNC", assert_C)
        return functionsmap[funcname](*args)

    @classmethod
    def evaluate_cell(cls, workbook, sheet, cell):
        column , row = extract_column_row(cell)
        
        e = cls(workbook)
        s = workbook[sheet]
        c = s[cell]

        return e.parse(c.value)

   
class ExpressionTreeBuilder(ExpressionEvaluator):


    # Grammar rules follow
    def expr(self):
        "expression ::= expr { ('='|'<'|'>'|'<='|'>=') expr }*"
        exprval = self.expr_()
        assert_ = lambda :self.nexttok.value in ["=",">","<","<=",">="]
        while self._accept('OPERATOR-INFIX',assert_):
            op = self.tok.value
            right = self.expr_()
            left = exprval
            exprval = (op, left, right)
        return exprval

    
    def expr_(self):
        "expression ::= term { ('+'|'-') term }*"
        exprval = self.term()
        assert_ = lambda :self.nexttok.value in ["+","-"]
        while self._accept('OPERATOR-INFIX',assert_):
            left = exprval
            op = self.tok.value
            right = self.term()
            left = exprval
            exprval = (op, left, right)
        return exprval

    def term(self):
        "term ::= power { ('*'|'/') power }*"
        termval = self.power()
        assert_ = lambda :self.nexttok.value in ["*","/"]
        while self._accept('OPERATOR-INFIX', assert_):
            left = termval
            op = self.tok.value
            right = self.power()
            if op == '*':
                termval = ("*",left, right)
            elif op == '/':
                termval =("/", left, right)
        return termval


    def power(self):
        "power ::= percent {^} percent"
        termval = self.percent()
        assert_ = lambda : self.nexttok.value == "^"
        while self._accept("OPERATOR-INFIX", assert_):
            op = self.tok.value
            right = self.percent()
            termval = (op, termval, right)
        return termval

    def percent(self):
        "percent ::= factor {%} "
        termval = self.factor()
        assert_ = lambda : self.nexttok.value == "%"
        while self._accept("OPERATOR-POSTFIX", assert_):
            op = self.tok.value
            termval = ("/", termval, 100)
        return termval
        
    
    def function(self):
        assert_C = lambda : self.nexttok.subtype == "CLOSE"
        assert_A = lambda : self.nexttok.subtype == "ARG"
    
        funcname,_ = self.tok.value.split("(")
        args = [self.expr()]
        while self._accept("SEP", assert_A):
            args.append(self.expr())
        self._expect("FUNC", assert_C)
        return (funcname, *args)


    def negation(self, pre):
        f = self.factor()
        return ("*", pre, f)
    
    def parsecell(self, c):
        if c.data_type == c.TYPE_FORMULA:
            e = ExpressionTreeBuilder(self.workbook)
            return e.parse(c.value)
        else: #c.data_type == c.TYPE_NUMERIC:
            return "!".join([self.workbook.get_active_sheet().title,c.coordinate])



if __name__ == "__main__":
    sys.setrecursionlimit(12000)
    filename = sys.argv[1]
    sheet = sys.argv[2]
    cell = sys.argv[3]
    w = load_workbook(filename)
    w.active = w.get_sheet_names().index(sheet)
    pretty_print(ExpressionTreeBuilder.evaluate_cell(w, sheet, cell))
    print(ExpressionEvaluator.evaluate_cell(w, sheet, cell))

